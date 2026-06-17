from __future__ import annotations

from pathlib import Path
import importlib.util

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
LEDGER_PATH = ROOT / "文档" / "原型逐页审校台账.xlsx"
BUILDER_PATH = ROOT / "scripts" / "build_handdrawn_menu_prototype_book.py"


HEADERS = [
    "页码",
    "编号",
    "模块",
    "菜单标题",
    "老系统菜单/功能",
    "Excel具体功能",
    "旧系统事实",
    "当前原型问题",
    "修正口径",
    "状态",
    "修正版号",
    "备注",
]


def load_builder():
    spec = importlib.util.spec_from_file_location("builder", BUILDER_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def build_rows():
    builder = load_builder()
    items = builder.merge_customer_master_items(builder.read_items())
    rows = []
    for page_no, item in enumerate(items, start=1):
        status = "锁定不动" if page_no <= 18 else "待审"
        note = "已审页，默认不主动重审" if page_no <= 18 else "从此页开始逐页审校"
        rows.append(
            [
                page_no,
                item["编号"],
                item["模块"],
                builder.client_title(item),
                item["老系统菜单/功能"],
                item["具体功能"],
                "",
                "",
                "",
                status,
                "",
                note,
            ]
        )
    return rows


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    locked_fill = PatternFill("solid", fgColor="F2F2F2")
    pending_fill = PatternFill("solid", fgColor="FFF2CC")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = center

    widths = {
        "A": 8,
        "B": 12,
        "C": 12,
        "D": 28,
        "E": 24,
        "F": 42,
        "G": 42,
        "H": 36,
        "I": 36,
        "J": 14,
        "K": 12,
        "L": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        page_cell = row[0]
        status_cell = row[9]
        for cell in row:
            cell.alignment = wrap
        page_cell.alignment = center
        status_cell.alignment = center
        if status_cell.value == "锁定不动":
            for cell in row:
                cell.fill = locked_fill
        elif status_cell.value == "待审":
            status_cell.fill = pending_fill


def main():
    LEDGER_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "逐页审校台账"

    for row_idx, row in enumerate(build_rows(), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    style_sheet(ws)
    wb.save(LEDGER_PATH)
    print(LEDGER_PATH)


if __name__ == "__main__":
    main()
