from __future__ import annotations

import math
from collections import Counter, defaultdict
from pathlib import Path
from re import split as re_split

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / "旅游接待管理系统-功能建设清单-完整保留版.xlsx"
OUT_DIR = ROOT / "output" / "甲方功能原型页面图"
INDEX_PATH = OUT_DIR / "00-原型页面图索引.png"

W, H = 1440, 1000
MARGIN = 42

COLORS = {
    "bg": "#f4f6fa",
    "panel": "#ffffff",
    "line": "#d8e0ea",
    "muted": "#64748b",
    "text": "#111827",
    "title": "#0f172a",
    "blue": "#2563eb",
    "blue_soft": "#e8f1ff",
    "green": "#16a34a",
    "green_soft": "#eafaf0",
    "red": "#dc2626",
    "red_soft": "#fee2e2",
    "orange": "#f59e0b",
    "orange_soft": "#fff7ed",
    "purple": "#7c3aed",
    "purple_soft": "#f3e8ff",
    "dark_panel": "#172033",
    "table_head": "#eef3f8",
}

MODULE_ORDER = ["客户管理", "采购管理", "销售管理", "计调操作", "财务管理", "数据统计", "企业资料", "系统设置"]
PHASE_COLOR = {"P0": COLORS["red"], "P1": COLORS["blue"], "P2": COLORS["purple"]}
PHASE_SOFT = {"P0": COLORS["red_soft"], "P1": COLORS["blue_soft"], "P2": COLORS["purple_soft"]}
PHASE_LABEL = {"P0": "P0 一期主线", "P1": "P1 增强/候选", "P2": "P2 后续规划"}

CASE = {
    "customer": "杭州远行国旅",
    "subject": "杭州远行国际旅行社有限公司",
    "team": "HZ20260518-003 西湖宋城二日游",
    "order": "S04-001",
    "amount": "¥86,400",
    "payable": "¥52,800",
    "guide": "陈晨",
    "vehicle": "浙A·D2398 37座",
    "hotel": "杭州湖滨酒店 18间",
    "supplier": "西湖车队 / 杭州湖滨酒店",
}


def clean(value: object) -> str:
    text = "" if value is None else str(value).strip()
    for old, new in {
        "帐款": "账款",
        "应收帐款": "应收账款",
        "应付帐款": "应付账款",
        "帐号": "账号",
        "帐户": "账户",
    }.items():
        text = text.replace(old, new)
    return text


def find_font(candidates: list[str], size: int) -> ImageFont.FreeTypeFont:
    for path in candidates:
        if Path(path).exists():
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


FONT_REG = find_font(
    [
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Light.ttc",
        "/Library/Fonts/Arial Unicode.ttf",
    ],
    20,
)


def font(size: int, bold: bool = False):
    if bold:
        return find_font(
            [
                "/System/Library/Fonts/PingFang.ttc",
                "/System/Library/Fonts/STHeiti Medium.ttc",
                "/Library/Fonts/Arial Unicode.ttf",
            ],
            size,
        )
    return find_font(
        [
            "/System/Library/Fonts/PingFang.ttc",
            "/System/Library/Fonts/STHeiti Light.ttc",
            "/Library/Fonts/Arial Unicode.ttf",
        ],
        size,
    )


F10 = font(10)
F12 = font(12)
F13 = font(13)
F14 = font(14)
F15 = font(15)
F16 = font(16)
F18 = font(18)
F20 = font(20)
F22 = font(22, True)
F26 = font(26, True)
F34 = font(34, True)


def read_items() -> list[dict[str, str]]:
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["功能建设清单"]
    headers = [clean(c.value) for c in ws[4]]
    items = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row[0]:
            continue
        items.append({headers[i]: clean(row[i]) for i in range(9)})
    return items


def split_features(text: str, limit: int = 8) -> list[str]:
    parts = [clean(x) for x in re_split(r"[、,，;；/]+", clean(text)) if clean(x)]
    result = []
    for item in parts:
        if item not in result:
            result.append(item)
    return result[:limit] or ["业务信息", "状态", "负责人", "附件"]


def classify(item: dict[str, str]) -> str:
    text = " ".join(item.values())
    module = item["模块"]
    title = item["优化建设内容"] or item["老系统菜单/功能"]
    menu = item["老系统菜单/功能"]

    if module == "数据统计":
        return "统计类"
    if module == "系统设置":
        return "系统设置类"
    if module == "企业资料":
        if any(k in title + menu for k in ["角色", "部门", "员工"]):
            return "系统设置类"
        return "档案类"
    if module == "客户管理":
        if "授信" in title or "应收" in title or "应收" in menu:
            return "财务类"
        if "合同" in title or "正式主体" in title or "授权" in title:
            return "订单团队类"
        return "档案类"
    if module == "采购管理":
        if any(k in title + menu for k in ["车辆", "排班"]):
            return "计调类"
        return "档案类"
    if module == "销售管理":
        if any(k in title + menu for k in ["费用变更", "成本分摊"]):
            return "财务类"
        if any(k in title + menu for k in ["智能报价", "票务", "门票"]):
            return "计调类"
        return "订单团队类"
    if module == "计调操作":
        if any(k in title + menu for k in ["报账"]):
            return "财务类"
        return "计调类"
    if module == "财务管理":
        return "财务类"
    if any(k in text for k in ["统计", "报表", "进度看板"]):
        return "统计类"
    if any(k in text for k in ["参数", "权限", "日志", "通知", "预警配置", "审批流"]):
        return "系统设置类"
    if any(k in text for k in ["应收", "应付", "收款", "付款", "发票", "备用金", "结算", "成本", "返佣"]):
        return "财务类"
    if item["模块"] == "计调操作" or any(k in text for k in ["计调", "导游", "车辆", "房态", "住宿", "车调", "派车", "报账"]):
        return "计调类"
    if any(k in text for k in ["订单", "团队", "团期", "产品", "拼团", "游客", "报价", "合同"]):
        return "订单团队类"
    return "档案类"


def rounded(draw: ImageDraw.ImageDraw, xy, radius=16, fill="#fff", outline=None, width=1):
    draw.rounded_rectangle(xy, radius=radius, fill=fill, outline=outline, width=width)


def text_size(draw: ImageDraw.ImageDraw, text: str, fnt) -> tuple[int, int]:
    box = draw.textbbox((0, 0), text, font=fnt)
    return box[2] - box[0], box[3] - box[1]


def draw_text(draw, xy, text, fnt, fill=COLORS["text"], max_width=None, line_gap=6, max_lines=None):
    x, y = xy
    if max_width is None:
        draw.text((x, y), text, font=fnt, fill=fill)
        return y + text_size(draw, text, fnt)[1]

    lines = wrap_text(draw, text, fnt, max_width)
    if max_lines:
        lines = lines[:max_lines]
    for line in lines:
        draw.text((x, y), line, font=fnt, fill=fill)
        y += text_size(draw, line, fnt)[1] + line_gap
    return y - line_gap


def wrap_text(draw, text, fnt, max_width) -> list[str]:
    text = clean(text)
    lines = []
    current = ""
    for ch in text:
        candidate = current + ch
        if text_size(draw, candidate, fnt)[0] <= max_width or not current:
            current = candidate
        else:
            lines.append(current)
            current = ch
    if current:
        lines.append(current)
    return lines


def pill(draw, x, y, text, fill, text_fill=None, fnt=F13):
    text_fill = text_fill or COLORS["text"]
    tw, th = text_size(draw, text, fnt)
    rounded(draw, (x, y, x + tw + 22, y + th + 12), radius=12, fill=fill)
    draw.text((x + 11, y + 6), text, font=fnt, fill=text_fill)
    return x + tw + 30


def button(draw, x, y, text, primary=False):
    fill = COLORS["blue"] if primary else COLORS["panel"]
    outline = COLORS["blue"] if primary else COLORS["line"]
    color = "#ffffff" if primary else COLORS["text"]
    tw, th = text_size(draw, text, F14)
    rounded(draw, (x, y, x + tw + 34, y + 36), radius=10, fill=fill, outline=outline)
    draw.text((x + 17, y + 9), text, font=F14, fill=color)
    return x + tw + 44


def card(draw, xy, title=None, fill=COLORS["panel"]):
    rounded(draw, xy, radius=18, fill=fill, outline=COLORS["line"])
    if title:
        x1, y1, x2, _ = xy
        draw.text((x1 + 18, y1 + 15), title, font=F18, fill=COLORS["title"])
        draw.line((x1, y1 + 54, x2, y1 + 54), fill=COLORS["line"], width=1)


def header(draw, item, idx):
    title = item["优化建设内容"] or item["老系统菜单/功能"]
    draw.text((MARGIN, 36), f"{idx:02d}. {item['模块']} / {title}", font=F34, fill=COLORS["title"])
    y = 88
    x = MARGIN
    x = pill(draw, x, y, item["编号"], PHASE_SOFT.get(item["优先级"], COLORS["blue_soft"]), PHASE_COLOR.get(item["优先级"], COLORS["blue"]), F14)
    x = pill(draw, x, y, PHASE_LABEL.get(item["优先级"], item["优先级"]), PHASE_SOFT.get(item["优先级"], COLORS["blue_soft"]), PHASE_COLOR.get(item["优先级"], COLORS["blue"]), F14)
    x = pill(draw, x, y, item["建设方式"], COLORS["green_soft"], COLORS["green"], F14)
    x = pill(draw, x, y, classify(item), COLORS["blue_soft"], COLORS["blue"], F14)
    draw_text(draw, (MARGIN, 130), f"菜单路径：{item['模块']} / {item['老系统菜单/功能']}", F16, COLORS["muted"], max_width=960, max_lines=1)


def page_usage_summary(item, kind):
    features = "、".join(split_features(item["具体功能"], 3))
    delivery = clean(item["交付说明"]).rstrip("。")
    if kind == "档案类":
        return f"维护{features}等基础信息，支撑下单、排团、结算和统计。{delivery}。"
    if kind == "订单团队类":
        return f"管理{features}等业务信息，承接销售接单、合同确认和团队流转。{delivery}。"
    if kind == "计调类":
        return f"安排{features}等计调资源，集中查看确认状态、资源冲突和成本变化。{delivery}。"
    if kind == "财务类":
        return f"管理{features}等财务事项，联动订单、团队、凭证、收付款和账款状态。{delivery}。"
    if kind == "统计类":
        return f"汇总{features}等经营数据，支持按客户、团队、资源和财务口径查看结果。{delivery}。"
    return f"配置{features}等系统规则，支撑权限、审批、预警、消息和日志留痕。{delivery}。"


def draw_filter_panel(draw, x, y, w, h, item, kind):
    card(draw, (x, y, x + w, y + h), "查询筛选")
    fields = {
        "档案类": ["名称/编号", "分类/状态", "负责人", "更新时间"],
        "订单团队类": ["客户", "团队日期", "订单状态", "销售负责人"],
        "计调类": ["团队号", "出团日期", "资源状态", "计调负责人"],
        "财务类": ["客户/供应商", "团队号", "账款状态", "账龄/日期"],
        "统计类": ["月份", "渠道/客户", "团队类型", "业务员"],
        "系统设置类": ["规则名称", "启停状态", "适用模块", "修改人"],
    }.get(kind, ["关键词", "状态", "负责人"])
    yy = y + 75
    for f in fields:
        rounded(draw, (x + 20, yy, x + w - 20, yy + 42), radius=10, fill=COLORS["bg"], outline=COLORS["line"])
        draw.text((x + 34, yy + 12), f, font=F14, fill=COLORS["muted"])
        yy += 54
    bx = x + 20
    bx = button(draw, bx, y + h - 58, "查询", True)
    button(draw, bx, y + h - 58, "重置", False)


def sample_rows(item, kind):
    title = item["优化建设内容"] or item["老系统菜单/功能"]
    if kind == "财务类":
        return [
            [CASE["team"], CASE["amount"], "待审核", "财务周敏"],
            ["费用变更 +¥4,800", "授信占用", "审批中", "销售张伟"],
            [CASE["supplier"], CASE["payable"], "待付款", "财务周敏"],
        ]
    if kind == "计调类":
        return [
            ["导游安排", CASE["guide"], "已确认", "无冲突"],
            ["车辆安排", CASE["vehicle"], "待比价", "车价偏高"],
            ["住宿安排", CASE["hotel"], "待确认", "房态紧张"],
        ]
    if kind == "统计类":
        return [
            ["本月散拼", "286人", "同比+18%", "已生成"],
            [CASE["team"], "毛利率33%", "¥28,600", "待复核"],
            [CASE["customer"], "账龄9天", CASE["amount"], "预警"],
        ]
    if kind == "系统设置类":
        return [
            ["超授信审批", "额度>100%", "财务→老板", "启用"],
            ["成本超预算", "超过5%", "计调→财务", "启用"],
            ["订单提醒", "确认件缺失", "销售负责人", "启用"],
        ]
    if kind == "订单团队类":
        return [
            [CASE["order"], CASE["customer"], CASE["amount"], "待确认"],
            ["拼团P20260518", "上海春秋门店", "32人", "已拼团"],
            [title[:10], CASE["team"][:14], "待补齐", "审批中"],
        ]
    return [
        [CASE["customer"], "A类客户", "授信80万", "正常"],
        [CASE["supplier"], "协议供应商", "月结", "有效"],
        [title[:12], "基础资料", "附件已传", "待确认"],
    ]


def draw_list_panel(draw, x, y, w, h, item, kind):
    card(draw, (x, y, x + w, y + h), "样例列表")
    headers = ["业务对象", "关联对象", "金额/指标", "状态"]
    col_w = [w * 0.34, w * 0.27, w * 0.2, w * 0.19]
    yy = y + 62
    draw.rectangle((x + 16, yy, x + w - 16, yy + 40), fill=COLORS["table_head"])
    xx = x + 26
    for i, htxt in enumerate(headers):
        draw.text((xx, yy + 11), htxt, font=F13, fill=COLORS["muted"])
        xx += col_w[i]
    yy += 48
    rows = sample_rows(item, kind)
    for r in rows:
        draw.line((x + 16, yy + 42, x + w - 16, yy + 42), fill=COLORS["line"], width=1)
        xx = x + 26
        for i, val in enumerate(r):
            draw_text(draw, (xx, yy + 10), val, F13, COLORS["text"], max_width=int(col_w[i] - 14), max_lines=1)
            xx += col_w[i]
        yy += 48
    bx = x + 20
    bx = button(draw, bx, y + h - 56, "详情", False)
    bx = button(draw, bx, y + h - 56, "编辑", False)
    button(draw, bx, y + h - 56, "提交审核", True)


def draw_detail_panel(draw, x, y, w, h, item, kind):
    card(draw, (x, y, x + w, y + h), "详情与状态")
    features = split_features(item["具体功能"], 6)
    info = detail_lines(kind, features)
    yy = y + 72
    for label, value in info:
        draw.text((x + 22, yy), label, font=F13, fill=COLORS["muted"])
        draw_text(draw, (x + 130, yy), value, F14, COLORS["text"], max_width=w - 160, max_lines=1)
        yy += 34
    draw.text((x + 22, yy + 10), "核心字段/操作", font=F16, fill=COLORS["title"])
    yy += 45
    for f in features:
        rounded(draw, (x + 22, yy, x + w - 22, yy + 32), radius=8, fill=COLORS["blue_soft"], outline=None)
        draw_text(draw, (x + 34, yy + 7), f, F13, COLORS["blue"], max_width=w - 68, max_lines=1)
        yy += 40


def detail_lines(kind, features):
    if kind == "财务类":
        return [("客户", CASE["customer"]), ("团队", CASE["team"]), ("应收", CASE["amount"]), ("应付", CASE["payable"]), ("凭证", "已上传2份")]
    if kind == "计调类":
        return [("团队", CASE["team"]), ("导游", CASE["guide"]), ("车辆", CASE["vehicle"]), ("住宿", CASE["hotel"]), ("异常", "房态紧张")]
    if kind == "统计类":
        return [("收客", "286人"), ("毛利", "54.8万"), ("欠款", "4户预警"), ("异常", "3个团队待处理"), ("口径", "成人/儿童计人数")]
    if kind == "系统设置类":
        return [("规则", features[0] if features else "审批规则"), ("范围", "销售/计调/财务"), ("节点", "主管→财务→老板"), ("状态", "启用"), ("日志", "保留")]
    if kind == "订单团队类":
        return [("客户", CASE["customer"]), ("团队", CASE["team"]), ("金额", CASE["amount"]), ("人数", "32人"), ("确认件", "待上传")]
    return [("名称", features[0] if features else "基础资料"), ("客户", CASE["customer"]), ("主体", CASE["subject"][:14]), ("状态", "正常"), ("附件", "已上传")]


def draw_side_panel(draw, x, y, w, h, item, kind):
    card(draw, (x, y, x + w, y + h), "页面能做什么")
    yy = y + 70
    status = status_lines(kind)
    for line in status:
        rounded(draw, (x + 22, yy, x + w - 22, yy + 40), radius=10, fill=COLORS["green_soft"], outline=None)
        draw_text(draw, (x + 36, yy + 11), line, F13, COLORS["green"], max_width=w - 68, max_lines=1)
        yy += 52
    yy += 10
    draw.text((x + 22, yy), "业务说明", font=F16, fill=COLORS["title"])
    yy += 34
    draw_text(draw, (x + 22, yy), business_explanation(item, kind), F14, COLORS["text"], max_width=w - 44, max_lines=6, line_gap=8)


def status_lines(kind):
    return {
        "档案类": ["启用/停用", "联动订单合同", "资料缺失提醒"],
        "订单团队类": ["待确认→已确认", "生成应收", "占用授信"],
        "计调类": ["待安排→可审核", "成本回写", "冲突提醒"],
        "财务类": ["待审→通过", "收付款联动", "账龄预警"],
        "统计类": ["自动汇总", "口径校验", "明细钻取"],
        "系统设置类": ["规则启停", "权限控制", "操作留痕"],
    }.get(kind, ["状态跟踪", "业务联动", "异常提醒"])


def business_explanation(item, kind):
    title = item["优化建设内容"] or item["老系统菜单/功能"]
    if kind == "档案类":
        return f"{title}是后续下单、排团、结算和统计的基础资料，先把主体、状态、负责人和附件维护清楚。"
    if kind == "订单团队类":
        return f"{title}承接客户需求和团队执行信息，确认后影响人数、应收、授信、计调安排和利润统计。"
    if kind == "计调类":
        return f"{title}让计调围绕一个团队集中安排资源，及时发现导游、车辆、房态、门票、成本和凭证异常。"
    if kind == "财务类":
        return f"{title}把订单收入、资源成本、凭证、收付款和结算状态关联起来，财务能提前看风险。"
    if kind == "统计类":
        return f"{title}从订单、团队、计调和财务数据自动汇总，让老板按统一口径看经营情况。"
    return f"{title}把关键规则配置化，减少靠人工记忆执行，保证审批、预警、权限和日志可追踪。"


def draw_flow_bar(draw, x, y, w, h, kind):
    steps = ["客户/合同", "销售订单", "计调排团", "财务审核", "经营统计"]
    active = {"档案类": 0, "订单团队类": 1, "计调类": 2, "财务类": 3, "统计类": 4, "系统设置类": 3}.get(kind, 1)
    gap = 10
    step_w = (w - gap * 4) / 5
    for i, s in enumerate(steps):
        xx = x + i * (step_w + gap)
        fill = COLORS["blue"] if i == active else COLORS["panel"]
        outline = COLORS["blue"] if i == active else COLORS["line"]
        color = "#fff" if i == active else COLORS["muted"]
        rounded(draw, (xx, y, xx + step_w, y + h), radius=12, fill=fill, outline=outline)
        tw, th = text_size(draw, s, F14)
        draw.text((xx + (step_w - tw) / 2, y + (h - th) / 2 - 1), s, font=F14, fill=color)


def generate_image(item, idx):
    kind = classify(item)
    img = Image.new("RGB", (W, H), COLORS["bg"])
    draw = ImageDraw.Draw(img)

    # Main canvas
    rounded(draw, (22, 22, W - 22, H - 22), radius=24, fill=COLORS["panel"], outline=COLORS["line"])
    header(draw, item, idx)
    draw_flow_bar(draw, MARGIN, 178, W - MARGIN * 2, 46, kind)

    top = 252
    draw_filter_panel(draw, MARGIN, top, 260, 548, item, kind)
    draw_list_panel(draw, MARGIN + 282, top, 530, 548, item, kind)
    draw_detail_panel(draw, MARGIN + 834, top, 315, 548, item, kind)
    draw_side_panel(draw, MARGIN + 1170, top, 186, 548, item, kind)

    # Footer delivery note
    y = 830
    rounded(draw, (MARGIN, y, W - MARGIN, H - 54), radius=16, fill=COLORS["orange_soft"], outline="#fed7aa")
    draw.text((MARGIN + 20, y + 16), "交付说明", font=F18, fill="#9a3412")
    draw_text(draw, (MARGIN + 20, y + 48), item["交付说明"], F15, COLORS["text"], max_width=W - MARGIN * 2 - 40, max_lines=3, line_gap=8)

    fname = f"{idx:02d}-{safe_name(item['模块'])}-{safe_name(item['优化建设内容'] or item['老系统菜单/功能'])}.png"
    out = OUT_DIR / fname
    img.save(out, quality=95)
    return out


def safe_name(text):
    text = clean(text)
    for ch in '/\\:*?"<>| ':
        text = text.replace(ch, "")
    return text[:28]


def generate_index(items, outputs):
    rows = []
    grouped = defaultdict(list)
    for idx, item in enumerate(items, start=1):
        grouped[item["模块"]].append((idx, item))
    for module in MODULE_ORDER:
        rows.append((module, len(grouped[module]), Counter(i["优先级"] for _, i in grouped[module])))

    img = Image.new("RGB", (W, H), COLORS["bg"])
    draw = ImageDraw.Draw(img)
    rounded(draw, (22, 22, W - 22, H - 22), radius=24, fill=COLORS["panel"], outline=COLORS["line"])
    draw.text((MARGIN, 46), "旅游接待管理系统 · 甲方功能原型页面图", font=F34, fill=COLORS["title"])
    draw_text(draw, (MARGIN, 100), "根据 Excel 功能建设清单自动生成。每张图片展示一个功能页面：筛选区、列表区、详情区、状态联动、核心功能和交付说明。", F18, COLORS["muted"], max_width=1180)
    pill(draw, MARGIN, 150, f"共 {len(items)} 张页面图", COLORS["blue_soft"], COLORS["blue"], F16)
    pill(draw, MARGIN + 190, 150, "不关联现有前端代码", COLORS["green_soft"], COLORS["green"], F16)

    x, y = MARGIN, 220
    card_w, card_h = 320, 140
    for idx, (module, total, counter) in enumerate(rows):
        col = idx % 4
        row = idx // 4
        xx = x + col * (card_w + 22)
        yy = y + row * (card_h + 26)
        rounded(draw, (xx, yy, xx + card_w, yy + card_h), radius=18, fill=COLORS["panel"], outline=COLORS["line"])
        draw.text((xx + 18, yy + 18), module, font=F22, fill=COLORS["title"])
        draw.text((xx + 18, yy + 56), f"{total} 张页面图", font=F16, fill=COLORS["muted"])
        px = xx + 18
        px = pill(draw, px, yy + 90, f"P0 {counter['P0']}", COLORS["red_soft"], COLORS["red"], F12)
        px = pill(draw, px, yy + 90, f"P1 {counter['P1']}", COLORS["blue_soft"], COLORS["blue"], F12)
        pill(draw, px, yy + 90, f"P2 {counter['P2']}", COLORS["purple_soft"], COLORS["purple"], F12)

    draw.text((MARGIN, 590), "建议给甲方看的顺序", font=F26, fill=COLORS["title"])
    advice = [
        "1. 先打开 00 索引页，告诉甲方这是一套功能页面图，不是最终 UI 视觉稿。",
        "2. 再按模块查看 67 张页面图，每张图对应 Excel 清单的一项功能。",
        "3. 重点看页面能做什么、有哪些字段、和一个团从接单到结算的关系。",
        "4. 如果甲方看不懂某项功能，就回到图片右侧“页面能做什么”和底部“交付说明”。",
    ]
    yy = 640
    for line in advice:
        draw_text(draw, (MARGIN, yy), line, F18, COLORS["text"], max_width=1180)
        yy += 44

    img.save(INDEX_PATH, quality=95)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for old in OUT_DIR.glob("*.png"):
        old.unlink()
    items = read_items()
    if len(items) != 67:
        raise RuntimeError(f"Expected 67 items, got {len(items)}")
    outputs = []
    for idx, item in enumerate(items, start=1):
        outputs.append(generate_image(item, idx))
    generate_index(items, outputs)
    print(OUT_DIR)
    print(len(outputs) + 1)


if __name__ == "__main__":
    main()
