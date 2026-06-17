from __future__ import annotations

import importlib.util
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
LEDGER_PATH = ROOT / "文档" / "原型逐页审校台账.xlsx"
REPORT_PATH = ROOT / "文档" / "图册整册业务逻辑审查报告-v093.md"
BUILD_SCRIPT_PATH = ROOT / "scripts" / "build_handdrawn_menu_prototype_book.py"
BASELINE_VERSION = "v093"


def load_build_module():
    spec = importlib.util.spec_from_file_location("prototype_book", BUILD_SCRIPT_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def build_excel_fact(item: dict[str, str]) -> str:
    parts = [
        f"Excel菜单/来源：{item.get('老系统菜单/功能', '')}",
        f"优化建设内容：{item.get('优化建设内容', '')}",
        f"具体功能：{item.get('具体功能', '')}",
        f"交付说明：{item.get('交付说明', '')}",
    ]
    return "；".join(part for part in parts if part and not part.endswith("："))


DEFAULT_PROBLEM = "未发现明显业务逻辑问题。当前页标题、页面作用与功能、页面操作与 Excel 主诉求基本一致。"
DEFAULT_FIX = "维持当前口径；后续如需调整，优先做术语精简、信息分层或页面可读性优化，不改主业务逻辑。"


ISSUE_OVERRIDES = {
    "C03": {
        "priority": "P1",
        "issue_type": "角色错位 / 关键规则缺失",
        "problem": "当前页虽然写到了授信、占用和超限审批，但整体仍偏通用财务列表页表达，客户维度、正式主体、占用释放链路和超限审批状态不够突出，容易被看成普通应收页。",
        "fix": "应把页面重心收回到“客户授信控制台”：突出正式主体、授信总额、已占用、可用额度、冻结额度、超限审批状态，并明确订单确认、费用变更、收款回款对额度的实时占用与释放。",
    },
    "C05": {
        "priority": "P1",
        "issue_type": "结构错位 / 关键规则缺失",
        "problem": "当前页业务方向基本正确，但仍偏通用合同台账表达，客户合同主体、到期提醒、下单合同校验和异常拦截的业务闭环还不够重，客户容易把它理解成普通附件台账。",
        "fix": "应强化合同主体、有效期、结款方式、到期提醒、下单校验和异常审批这条主链路；PNG 结构要明确“合同台账 + 下单校验结果 + 到期/无效状态”。",
    },
    "S02": {
        "priority": "P1",
        "issue_type": "角色边界不清",
        "problem": "当前页已经按你的口径改成销售侧团队/团期总入口，但与后面的“团队主档”边界仍不够清楚，客户容易疑惑两个页面是否重复。",
        "fix": "应继续明确：本页负责按成团方式切换、筛选、建团和查看资源入口；团队主档负责单团详情、内部说明、团内订单和团队级操作。标题不改，但文案和截图应进一步拉开职责。",
    },
    "S02-AUTO": {
        "priority": "P0",
        "issue_type": "关键逻辑缺失 / 信息过载",
        "problem": "当前长图把产品带入、资源安排、团队操作台、单据打印、用餐/自费/购物/地接/附加等全部堆在一页里，虽然覆盖面大，但“自动带入什么、人工确认什么、保存后回写什么”的主链路还不够清晰，难以作为甲方确认自动化逻辑的依据。",
        "fix": "应保留这页作为总逻辑图，但把主线强制压成三段：产品默认参数带入、计调逐项确认、成本/审核/单据回写；其它细节只保留关键分支，团队操作台按钮继续交给后续按钮页承接，不在本页重复堆叠。",
    },
    "S03": {
        "priority": "P1",
        "issue_type": "角色边界不清",
        "problem": "团队主档页当前承担了团队详情总入口的职责，但和团期管理页之间的职责边界仍不够直观，容易出现“都是团队入口”的重复感。",
        "fix": "应明确本页只负责单团队详情、阶段信息、产品说明、收客须知、内部备注和团内订单；新增、切换成团方式、筛选和资源入口留在团期管理页。",
    },
    "S16": {
        "priority": "P1",
        "issue_type": "能力表述过满 / 人工边界不清",
        "problem": "当前页把企业微信群接入、需求识别、自动回复、自动生成工单都写得比较满，但人工接管、失败回退、工单归属和跟进责任边界不够清晰，容易让客户误解成全自动销售机器人。",
        "fix": "应把定位收成“群消息辅助分发和工单生成”：明确自动识别只是辅助，复杂需求必须人工接管；需要补齐工单归属人、接管节点、遗漏提醒和处理状态闭环。",
    },
    "S17": {
        "priority": "P1",
        "issue_type": "关键规则缺失 / 能力表述过满",
        "problem": "当前页强调智能生成行程和报价，但报价规则、价格来源、人工微调和最终确认边界不够清楚，容易让客户理解为系统可以直接给出可成交报价。",
        "fix": "应明确它是“智能生成草案”而非最终报价：突出资料库匹配、模块化拼装、资源价格来源、人工调整、生成标准文件和最终确认流程，避免把 AI 能力写成全自动成交系统。",
    },
    "D01": {
        "priority": "P1",
        "issue_type": "页面重心偏移",
        "problem": "团队安排总控台已经并入酒店和车辆两组能力，但当前页仍是较均匀的多资源总控表达，没有完全体现“计调主要先安排房子和车子”的业务优先级。",
        "fix": "应把住/车入口和状态放成主视觉，其它资源作为次级入口；明确总控台的第一任务是看哪些团队还没落实房和车，再分流到酒店房态页和派车页。",
    },
}


CONFIRM_NOTES = {
    "C01": "当前页已吸收 Excel 中的 C02、C04；业务逻辑基本成立，但需求源与图册结构需要后续同步。",
    "P02": "当前页已按后续业务澄清收去“协议/结算方式”主表达，业务逻辑自洽；建议后续回填 Excel。",
    "P03": "当前页已按后续业务澄清收去固定采购价表达，改成资源可供关系页；业务逻辑自洽，建议同步 Excel。",
    "P04": "当前页已按后续口径收成采购合同台账与校验页，不再强调新增上传；建议同步 Excel 的旧字段描述。",
    "S06": "当前页标题已按业务收口为“拼团订单”，比 Excel 的“多订单聚合”更贴近客户理解，建议后续回填需求文档。",
    "F01-IMPREST": "该页为补充页，当前业务逻辑明确，但它属于图册新增页，后续建议回填到 Excel 需求源。",
    "F05": "页名已从“应收明细来源追踪”收短为“应收明细”；来源追溯保留为页面能力，建议同步 Excel 标题。",
    "F07": "页名已从“应付明细来源追踪”收短为“应付明细”；来源追溯保留为页面能力，建议同步 Excel 标题。",
    "F12": "当前页已按业务收口为“收款记录”，确认件匹配作为辅助能力保留；建议同步 Excel 标题和口径。",
    "F09": "当前页已与团队预付款拆分表达，导游备用金闭环逻辑清楚；建议后续把 Excel 的混合口径同步拆开。",
}


MERGED_OR_REMOVED = {
    "P05": {
        "page": "并入29/31",
        "status": "图册已并入 D05/D09",
        "problem": "该功能当前不再作为独立图册页展示，车辆资源与日期排班能力已吸收到计调侧酒店/车辆主页面体系中。",
        "fix": "如维持当前图册结构，应在 D09 车调询价与派车系统中继续承接车辆资源池、日期占用和团队派车逻辑；同时同步需求文档。",
        "remark": "并入 D09 / 历史独立页已取消",
    },
    "P05-CAL": {
        "page": "并入31",
        "status": "图册已并入 D05/D09",
        "problem": "该功能当前不再作为独立图册页展示，车辆日历能力已吸收到车辆派车主页面中。",
        "fix": "如维持当前图册结构，应在 D09 中继续承接日期占用、空闲/占用/维修状态和团队占用信息；同时同步需求文档。",
        "remark": "并入 D09 / 历史独立页已取消",
    },
    "P06": {
        "page": "并入29",
        "status": "图册已并入 D05/D09",
        "problem": "该功能当前不再作为独立图册页展示，酒店主档能力已并入计调侧酒店房态页。",
        "fix": "如维持当前图册结构，应在 D05 中继续同时表达酒店主档、日期房态和房调动作；同时同步需求文档。",
        "remark": "并入 D05 / 历史独立页已取消",
    },
    "P07": {
        "page": "并入29",
        "status": "图册已并入 D05/D09",
        "problem": "该功能当前不再作为独立图册页展示，房调页已并入酒店房态主页面。",
        "fix": "如维持当前图册结构，应在 D05 中继续承接占房、调房、加房、减房、锁房和释放逻辑；同时同步需求文档。",
        "remark": "并入 D05 / 历史独立页已取消",
    },
    "S07": {
        "page": "已删除",
        "status": "图册已删除",
        "problem": "Excel 中的“多订单共车成本分摊”当前不再保留独立图册页，相关逻辑已收回到团队资源安排自动化逻辑中表达。",
        "fix": "如保持当前口径，应在 S02-AUTO 中说明共车和成本均摊只是团队安排中的一类处理逻辑，不再单独成页；同时同步需求文档。",
        "remark": "已删除并并入 S02-AUTO",
    },
}


def build_review_payload(code: str, item: dict[str, str]):
    base = {
        "status": "已确认",
        "problem": DEFAULT_PROBLEM,
        "fix": DEFAULT_FIX,
        "priority": "P2",
        "issue_type": "未发现明显业务逻辑问题",
        "remark": "审查基线 v093",
    }
    if code in ISSUE_OVERRIDES:
        override = ISSUE_OVERRIDES[code]
        base.update(
            {
                "status": "待修正",
                "problem": override["problem"],
                "fix": override["fix"],
                "priority": override["priority"],
                "issue_type": override["issue_type"],
                "remark": "审查基线 v093",
            }
        )
    note = CONFIRM_NOTES.get(code)
    if note:
        base["remark"] = note
    return base


def update_ledger():
    mod = load_build_module()
    source_items = mod.read_items()
    source_map = {item["编号"]: item for item in source_items}
    current_items = mod.merge_customer_master_items(source_items)
    current_map = {}
    for idx, item in enumerate(current_items, start=1):
        current_map[item["编号"]] = {
            "page": idx,
            "module": item["模块"],
            "title": mod.client_title(item),
            "item": item,
            "usage": mod.page_usage(item),
            "operations": mod.page_operations(item),
        }

    wb = load_workbook(LEDGER_PATH)
    ws = wb.active

    ledger_rows = {}
    for row_idx in range(2, ws.max_row + 1):
        code = ws.cell(row_idx, 2).value
        if code:
            ledger_rows[str(code).strip()] = row_idx

    report_entries = []

    for code, info in current_map.items():
        row_idx = ledger_rows.get(code)
        if row_idx is None:
            continue
        item = info["item"]
        payload = build_review_payload(code, item)
        ws.cell(row_idx, 1).value = info["page"]
        ws.cell(row_idx, 3).value = info["module"]
        ws.cell(row_idx, 4).value = info["title"]
        ws.cell(row_idx, 5).value = item.get("老系统菜单/功能", "")
        ws.cell(row_idx, 6).value = item.get("具体功能", "")
        ws.cell(row_idx, 7).value = build_excel_fact(source_map.get(code, item))
        ws.cell(row_idx, 8).value = payload["problem"]
        ws.cell(row_idx, 9).value = payload["fix"]
        ws.cell(row_idx, 10).value = payload["status"]
        ws.cell(row_idx, 11).value = BASELINE_VERSION
        ws.cell(row_idx, 12).value = f"{payload['priority']}｜{payload['issue_type']}｜{payload['remark']}"
        report_entries.append(
            {
                "page": info["page"],
                "code": code,
                "module": info["module"],
                "title": info["title"],
                "status": payload["status"],
                "priority": payload["priority"],
                "issue_type": payload["issue_type"],
                "problem": payload["problem"],
                "fix": payload["fix"],
                "usage": info["usage"],
                "operations": info["operations"],
                "remark": payload["remark"],
            }
        )

    for code, merged in MERGED_OR_REMOVED.items():
        row_idx = ledger_rows.get(code)
        if row_idx is None:
            continue
        ws.cell(row_idx, 1).value = merged["page"]
        ws.cell(row_idx, 7).value = "Excel 功能仍存在，但当前 atlas 已调整为合并或删除口径。"
        ws.cell(row_idx, 8).value = merged["problem"]
        ws.cell(row_idx, 9).value = merged["fix"]
        ws.cell(row_idx, 10).value = merged["status"]
        ws.cell(row_idx, 11).value = BASELINE_VERSION
        ws.cell(row_idx, 12).value = merged["remark"]

    wb.save(LEDGER_PATH)

    write_report(mod, source_map, current_map, report_entries)


def write_report(mod, source_map, current_map, report_entries):
    report_entries = sorted(report_entries, key=lambda x: x["page"])
    current_codes = set(current_map)
    source_codes = set(source_map)
    source_only = sorted(source_codes - current_codes)
    current_only = sorted(current_codes - source_codes)

    merged_notes = {
        "C02": "已并入 01 客户主档",
        "C04": "已并入 01 客户主档",
        "C06": "当前图册未单独成页，属于需求源与图册结构未同步项",
        "P05": "已并入 31 车调询价与派车系统",
        "P06": "已并入 29 酒店房间与房态库存管理",
        "P08": "当前图册未单独成页，属于需求源与图册结构未同步项",
        "P09": "当前图册未单独成页，属于需求源与图册结构未同步项",
        "S07": "当前图册已删除并并入 11 团期资源安排自动化逻辑图",
    }
    supplement_notes = {
        "S01-FORM": "产品资料补充页",
        "S01-RESOURCE": "产品资料补充页",
        "S02-AUTO": "团期管理补充页",
        "S02-TOOLS": "团队安排补充页",
        "S02-OPS-DETAIL": "团队操作台补充页",
        "S02-OPS-RISK": "团队操作台补充页",
        "S04-RULES": "订单管理补充页",
        "F01-IMPREST": "财务补充页",
    }

    status_counter = Counter(entry["status"] for entry in report_entries)
    priority_counter = Counter(entry["priority"] for entry in report_entries if entry["status"] == "待修正")
    issue_entries = [entry for entry in report_entries if entry["status"] == "待修正"]

    lines = [
        f"# 图册整册业务逻辑审查报告-{BASELINE_VERSION}",
        "",
        "## 审查基线",
        f"- 主依据：`旅游接待管理系统-功能建设清单-完整保留版.xlsx`",
        f"- 图册基线：`旅游接待管理系统-甲方菜单页面原型图册.docx`（{BASELINE_VERSION}）",
        "- 审查原则：Excel 定稿需求优先；当前 Word/PNG 是否一致；旧系统仅作辅助参考",
        "- 本轮只做整册业务逻辑审查，不改图册脚本、不重生 Word/PNG",
        "",
        "## 总体结论",
        f"- 当前 atlas 功能页：`{len(report_entries)}` 页",
        f"- 本轮判定 `已确认`：`{status_counter.get('已确认', 0)}` 页",
        f"- 本轮判定 `待修正`：`{status_counter.get('待修正', 0)}` 页",
        f"- 待修正优先级分布：`P0={priority_counter.get('P0', 0)}`，`P1={priority_counter.get('P1', 0)}`",
        "- 当前最大风险不在数量，而在几类关键页：授信/合同、团期与团队入口边界、团期资源安排自动化逻辑、团队安排总控台、AI 销售辅助页。",
        "",
        "## 需求源与图册结构差异",
        "### Excel 有但当前未单独成页",
    ]
    for code in source_only:
        src = source_map[code]
        lines.append(f"- `{code} {src['模块']} / {src['优化建设内容']}`：{merged_notes.get(code, '未单独成页，需要后续同步需求源与图册结构。')}")

    lines.extend(
        [
            "",
            "### 当前图册补充页（Excel 原始清单中无独立编号）",
        ]
    )
    for code in current_only:
        item = current_map[code]["item"]
        lines.append(f"- `{code} {item['模块']} / {mod.client_title(item)}`：{supplement_notes.get(code, '当前 atlas 的补充说明页。')}")

    lines.extend(
        [
            "",
            "### 标题已按业务收口、建议后续同步 Excel 的页面",
            "- `18 拼团订单`：Excel 原标题仍是“拼团订单与多订单聚合”",
            "- `36 应收明细`：Excel 原标题仍是“应收明细来源追踪”",
            "- `38 应付明细`：Excel 原标题仍是“应付明细来源追踪”",
            "- `43 收款记录`：Excel 原标题仍是“收款与确认件匹配”",
            "",
            "## 高优先级待修正页面",
            "| 页码 | 编号 | 页面 | 优先级 | 问题类型 | 问题摘要 |",
            "| --- | --- | --- | --- | --- | --- |",
        ]
    )
    for entry in issue_entries:
        lines.append(
            f"| {entry['page']:02d} | {entry['code']} | {entry['module']} / {entry['title']} | {entry['priority']} | {entry['issue_type']} | {entry['problem']} |"
        )

    lines.extend(
        [
            "",
            "## 逐页审查结论",
        ]
    )
    for entry in report_entries:
        lines.extend(
            [
                f"### {entry['page']:02d}. {entry['module']} / {entry['title']}",
                f"- 状态：`{entry['status']}`",
                f"- Excel事实：{build_excel_fact(source_map.get(entry['code'], current_map[entry['code']]['item']))}",
                f"- 当前原型问题：{entry['problem']}",
                f"- 修正口径：{entry['fix']}",
                f"- 备注：{entry['priority']}｜{entry['issue_type']}｜{entry['remark']}",
                "",
            ]
        )

    lines.extend(
        [
            "## 整改分批建议",
            "1. 第一批先改 `C03`、`C05`、`S02`、`S03`、`S02-AUTO`、`D01`：这些页直接影响客户对主业务链路的理解。",
            "2. 第二批改 `S16`、`S17`：避免 AI 页能力表达过满、承诺过重。",
            "3. 第三批同步需求源：把 Excel 与 atlas 当前合并/删页/补页口径回填，减少后续再次审查时的源头冲突。",
            "",
        ]
    )

    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    update_ledger()
    print(LEDGER_PATH)
    print(REPORT_PATH)
